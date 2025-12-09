"""
File content extraction module supporting various file types.
"""
import os
import json
import mimetypes
from pathlib import Path
from typing import Optional, Dict, Any
import chardet

# Document processing
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from PIL import Image
    from PIL.ExifTags import TAGS
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class FileExtractor:
    """Extract content and metadata from various file types."""
    
    def __init__(self, max_text_size: int = 50000, transcribe_videos: bool = False, video_transcriber=None):
        """
        Initialize the file extractor.
        
        Args:
            max_text_size: Maximum characters to extract from text files
            transcribe_videos: Whether to transcribe video/audio files
            video_transcriber: VideoTranscriber instance (if transcribing)
        """
        self.max_text_size = max_text_size
        self.transcribe_videos = transcribe_videos
        self.video_transcriber = video_transcriber
        
    def extract(self, file_path: Path) -> Dict[str, Any]:
        """
        Extract content and metadata from a file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Dictionary containing file metadata and extracted content
        """
        result = {
            'path': str(file_path),
            'name': file_path.name,
            'extension': file_path.suffix.lower(),
            'size': file_path.stat().st_size,
            'content': '',
            'metadata': {},
            'type': 'unknown'
        }
        
        try:
            # Determine file type
            mime_type, _ = mimetypes.guess_type(str(file_path))
            result['mime_type'] = mime_type or 'unknown'
            
            # Extract based on file type
            ext = result['extension']
            
            if ext in ['.txt', '.md', '.rst', '.log']:
                result['type'] = 'text'
                result['content'] = self._extract_text(file_path)
            elif ext == '.pdf' and PDF_AVAILABLE:
                result['type'] = 'pdf'
                result['content'] = self._extract_pdf(file_path)
            elif ext in ['.docx', '.doc'] and DOCX_AVAILABLE:
                result['type'] = 'document'
                result['content'] = self._extract_docx(file_path)
            elif ext in ['.json', '.jsonl']:
                result['type'] = 'json'
                result['content'] = self._extract_json(file_path)
            elif ext in ['.py', '.js', '.java', '.cpp', '.c', '.h', '.hpp', '.cs', '.go', '.rs', '.rb', '.php', '.swift', '.kt', '.ts', '.tsx', '.jsx']:
                result['type'] = 'code'
                result['content'] = self._extract_text(file_path)
            elif ext in ['.html', '.htm', '.xml', '.svg']:
                result['type'] = 'markup'
                result['content'] = self._extract_text(file_path)
            elif ext in ['.csv', '.tsv']:
                result['type'] = 'data'
                result['content'] = self._extract_text(file_path)
            elif ext in ['.xlsx', '.xls'] and EXCEL_AVAILABLE:
                result['type'] = 'spreadsheet'
                result['content'] = self._extract_excel(file_path)
            elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.tif'] and PIL_AVAILABLE:
                result['type'] = 'image'
                result['metadata'] = self._extract_image_metadata(file_path)
                result['content'] = f"Image file: {file_path.name}, {result['metadata'].get('dimensions', 'unknown dimensions')}"
            elif ext in ['.mp4', '.avi', '.mov', '.mkv', '.wmv', '.flv', '.webm', '.m4v']:
                result['type'] = 'video'
                if self.transcribe_videos and self.video_transcriber:
                    transcription = self.video_transcriber.transcribe_video(file_path)
                    if transcription['success']:
                        result['content'] = f"Video file: {file_path.name}\n\nTranscription:\n{transcription['text']}"
                        result['transcription'] = transcription['text']
                    else:
                        result['content'] = f"Video file: {file_path.name} (transcription failed)"
                else:
                    result['content'] = f"Video file: {file_path.name}"
            elif ext in ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.m4a', '.wma']:
                result['type'] = 'audio'
                if self.transcribe_videos and self.video_transcriber:
                    transcription = self.video_transcriber.transcribe_audio(file_path)
                    if transcription['success']:
                        result['content'] = f"Audio file: {file_path.name}\n\nTranscription:\n{transcription['text']}"
                        result['transcription'] = transcription['text']
                    else:
                        result['content'] = f"Audio file: {file_path.name} (transcription failed)"
                else:
                    result['content'] = f"Audio file: {file_path.name}"
            elif ext in ['.zip', '.tar', '.gz', '.rar', '.7z', '.bz2']:
                result['type'] = 'archive'
                result['content'] = f"Archive file: {file_path.name}"
            else:
                # Try to read as text if small enough
                if result['size'] < 1024 * 1024:  # 1MB
                    try:
                        result['content'] = self._extract_text(file_path)
                        result['type'] = 'text'
                    except:
                        result['content'] = f"Binary or unknown file: {file_path.name}"
                else:
                    result['content'] = f"Large binary file: {file_path.name}"
                    
        except Exception as e:
            result['content'] = f"Error extracting {file_path.name}: {str(e)}"
            result['error'] = str(e)
            
        return result
    
    def _extract_text(self, file_path: Path) -> str:
        """Extract text from plain text files."""
        try:
            # Detect encoding
            with open(file_path, 'rb') as f:
                raw_data = f.read(min(100000, file_path.stat().st_size))
                detected = chardet.detect(raw_data)
                encoding = detected['encoding'] or 'utf-8'
            
            # Read file
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                content = f.read(self.max_text_size)
                
            return content
        except Exception as e:
            return f"Error reading text file: {str(e)}"
    
    def _extract_pdf(self, file_path: Path) -> str:
        """Extract text from PDF files."""
        try:
            reader = PdfReader(str(file_path))
            text_parts = []
            
            # Extract from first few pages
            for page_num, page in enumerate(reader.pages[:10]):  # First 10 pages
                text_parts.append(page.extract_text())
                if len(''.join(text_parts)) > self.max_text_size:
                    break
                    
            return '\n'.join(text_parts)[:self.max_text_size]
        except Exception as e:
            return f"Error reading PDF: {str(e)}"
    
    def _extract_docx(self, file_path: Path) -> str:
        """Extract text from DOCX files."""
        try:
            doc = Document(str(file_path))
            paragraphs = [p.text for p in doc.paragraphs]
            return '\n'.join(paragraphs)[:self.max_text_size]
        except Exception as e:
            return f"Error reading DOCX: {str(e)}"
    
    def _extract_json(self, file_path: Path) -> str:
        """Extract and format JSON content."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            formatted = json.dumps(data, indent=2)
            return formatted[:self.max_text_size]
        except Exception as e:
            # Try reading as text if JSON parsing fails
            return self._extract_text(file_path)
    
    def _extract_excel(self, file_path: Path) -> str:
        """Extract text from Excel files."""
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            text_parts = []
            
            for sheet_name in wb.sheetnames[:5]:  # First 5 sheets
                sheet = wb[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}")
                
                for row_idx, row in enumerate(sheet.iter_rows(max_row=100, values_only=True)):
                    row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                    text_parts.append(row_text)
                    
                    if len('\n'.join(text_parts)) > self.max_text_size:
                        break
                        
            return '\n'.join(text_parts)[:self.max_text_size]
        except Exception as e:
            return f"Error reading Excel file: {str(e)}"
    
    def _extract_image_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract metadata from image files."""
        try:
            with Image.open(file_path) as img:
                metadata = {
                    'dimensions': f"{img.width}x{img.height}",
                    'format': img.format,
                    'mode': img.mode
                }
                
                # Extract EXIF data if available
                exif_data = img.getexif()
                if exif_data:
                    exif = {}
                    for tag_id, value in exif_data.items():
                        tag = TAGS.get(tag_id, tag_id)
                        exif[tag] = str(value)[:100]  # Limit value length
                    metadata['exif'] = exif
                    
                return metadata
        except Exception as e:
            return {'error': str(e)}
